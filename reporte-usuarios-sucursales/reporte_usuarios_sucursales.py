"""
Reporte: Usuarios y Sucursales — Naturisa
Genera un Excel con dos hojas:
  - Por Sucursal: una fila por relacion usuario-sucursal, agrupado por sucursal
  - Por Usuario:  una fila por usuario con todas sus sucursales

USO:
  python reporte_usuarios_sucursales.py

El token se obtiene desde la app web de Naturisa (F12 → Network → cualquier
request → Header: Authorization). Pegarlo en la variable TOKEN de abajo.
"""

import re, sys, requests
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── CONFIGURACION ─────────────────────────────────────────────────────────────
TOKEN  = "PEGAR_TOKEN_AQUI"   # sin el prefijo "Bearer"
OUTPUT = rf"C:\Scripts\Naturisa\reporte-usuarios-sucursales\Usuarios_Sucursales_{date.today().strftime('%Y%m%d')}.xlsx"
# ─────────────────────────────────────────────────────────────────────────────

BASE    = "https://gateway.naturisa.com.ec/bff/web/security/backoffice/api"
HEADERS = {"Authorization": f"Bearer {TOKEN}", "Accept": "application/json"}

def get(url):
    r = requests.get(url, headers=HEADERS, timeout=15)
    r.raise_for_status()
    return r.json()

# ── 1. Usuarios activos ───────────────────────────────────────────────────────
print("Cargando usuarios...", flush=True)
all_users = get(f"{BASE}/users?pageSize=5000&status=ACTIVO")["data"]["data"]
print(f"  {len(all_users)} usuarios", flush=True)

# ── 2. Warehouses → mapa subsidiaryId → nombre limpio ────────────────────────
print("Cargando sucursales...", flush=True)
wh_data = get(
    "https://gateway.naturisa.com.ec/bff/web/ap1/backoffice/api/warehouses"
    "?warehouseTypeCodes=21&warehouseTypeCodes=22&warehouseTypeCodes=23"
    "&status=ACTIVO&pageSize=1000&orderBy=name"
)["data"]["data"] or []

subs_map = {}
for w in wh_data:
    sid = w["subsidiaryId"]
    if sid not in subs_map or w["name"].startswith("BF "):
        subs_map[sid] = re.sub(r"^(BA|BF|BM)\s+", "", w["name"]).strip()
print(f"  {len(subs_map)} sucursales mapeadas", flush=True)

# ── 3. Sucursales por usuario en paralelo ────────────────────────────────────
def fetch_user_subs(u):
    uid = u["idUser"]
    try:
        items = get(f"{BASE}/user_subsidiaries?pageSize=500&userId={uid}&status=ACTIVO")["data"]["data"] or []
        names = sorted(subs_map.get(it["subsidiaryId"], f"Sub {it['subsidiaryId']}") for it in items)
        return uid, names
    except Exception:
        return uid, []

print(f"Consultando {len(all_users)} usuarios en paralelo...", flush=True)
uid_to_subs = {}
done = 0
with ThreadPoolExecutor(max_workers=20) as pool:
    futures = {pool.submit(fetch_user_subs, u): u for u in all_users}
    for fut in as_completed(futures):
        uid, names = fut.result()
        uid_to_subs[uid] = names
        done += 1
        if done % 200 == 0:
            print(f"  {done}/{len(all_users)}...", flush=True)

sin_sub = sum(1 for s in uid_to_subs.values() if not s)
print(f"  Con sucursales: {len(all_users) - sin_sub}  /  Sin sucursales: {sin_sub}", flush=True)

# ── 4. Preparar datos ─────────────────────────────────────────────────────────
# Hoja 1: una fila por relacion usuario-sucursal
rows_by_sub = []
for u in all_users:
    uid  = u["idUser"]
    info = {"ID": uid, "Nombre": f"{u['firstNames']} {u['lastNames']}",
            "Usuario": u["username"], "Email": u.get("email", "")}
    for sub in uid_to_subs.get(uid, []):
        rows_by_sub.append({**info, "Sucursal": sub})
rows_by_sub.sort(key=lambda r: (r["Sucursal"], r["Nombre"]))

# Hoja 2: una fila por usuario con sucursales agregadas
user_subs_agg = defaultdict(list)
for r in rows_by_sub:
    user_subs_agg[r["ID"]].append(r["Sucursal"])

rows_by_user = []
for u in all_users:
    uid  = u["idUser"]
    subs = user_subs_agg.get(uid, [])
    rows_by_user.append({
        "ID":      uid,
        "Nombre":  f"{u['firstNames']} {u['lastNames']}",
        "Usuario": u["username"],
        "Email":   u.get("email", ""),
        "Sucursales":      ", ".join(sorted(subs)),
        "TotalSucursales": len(subs),
    })
rows_by_user.sort(key=lambda r: r["Nombre"])

# ── 5. Estilos ────────────────────────────────────────────────────────────────
hdr_fill  = PatternFill("solid", fgColor="1B3A6B")
hdr_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
alt_fill  = PatternFill("solid", fgColor="EEF2FA")
warn_fill = PatternFill("solid", fgColor="FFF3CD")
thin      = Side(style="thin", color="D0D0D0")
brd       = Border(left=thin, right=thin, top=thin, bottom=thin)

COLORS = [
    "DDEEFF","D5F5E3","FCF3CF","FADBD8","E8DAEF","D6EAF8","FDEBD0",
    "EBF5FB","F9EBEA","E9F7EF","FEF9E7","F4ECF7","EAF2F8","FDF2E9",
    "E8F8F5","FEF5E7","F5EEF8","E8F6F3","F0F3FF","FFF8E7","F0FFF0",
    "FFF0F0","F0F0FF","FFFEF0","F0FFFF","FDEBD0","D6EAF8",
]
sub_names = sorted(set(r["Sucursal"] for r in rows_by_sub))
sub_color = {s: PatternFill("solid", fgColor=COLORS[i % len(COLORS)])
             for i, s in enumerate(sub_names)}

def write_header(ws, headers):
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(1, col)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = brd
    ws.row_dimensions[1].height = 22

def set_cell(ws, row, col, fill=None, wrap=False):
    cell = ws.cell(row, col)
    cell.border = brd
    cell.alignment = Alignment(vertical="center", wrap_text=wrap)
    if fill:
        cell.fill = fill

# ── 6. Generar Excel ──────────────────────────────────────────────────────────
print("Generando Excel...", flush=True)
wb = Workbook()

# Hoja 1 — Por Sucursal
ws1 = wb.active
ws1.title = "Por Sucursal"
write_header(ws1, ["Sucursal", "ID", "Nombre", "Usuario", "Email"])
for i, row in enumerate(rows_by_sub, 2):
    ws1.append([row["Sucursal"], row["ID"], row["Nombre"], row["Usuario"], row["Email"]])
    fill = sub_color.get(row["Sucursal"])
    for col in range(1, 6):
        set_cell(ws1, i, col, fill)
ws1.column_dimensions["A"].width = 28
ws1.column_dimensions["B"].width = 9
ws1.column_dimensions["C"].width = 32
ws1.column_dimensions["D"].width = 18
ws1.column_dimensions["E"].width = 36
ws1.freeze_panes = "A2"
ws1.auto_filter.ref = ws1.dimensions

# Hoja 2 — Por Usuario
ws2 = wb.create_sheet("Por Usuario")
write_header(ws2, ["ID", "Nombre", "Usuario", "Email", "Sucursales", "# Sucursales"])
for i, row in enumerate(rows_by_user, 2):
    ws2.append([row["ID"], row["Nombre"], row["Usuario"], row["Email"],
                row["Sucursales"], row["TotalSucursales"]])
    fill = warn_fill if row["TotalSucursales"] == 0 else (alt_fill if i % 2 == 0 else None)
    for col in range(1, 7):
        set_cell(ws2, i, col, fill, wrap=(col == 5))
ws2.column_dimensions["A"].width = 9
ws2.column_dimensions["B"].width = 32
ws2.column_dimensions["C"].width = 18
ws2.column_dimensions["D"].width = 36
ws2.column_dimensions["E"].width = 65
ws2.column_dimensions["F"].width = 14
ws2.freeze_panes = "A2"
ws2.auto_filter.ref = ws2.dimensions

wb.save(OUTPUT)
print(f"\nArchivo guardado: {OUTPUT}")
print(f"  Hoja 'Por Sucursal': {len(rows_by_sub)} filas")
print(f"  Hoja 'Por Usuario':  {len(rows_by_user)} filas ({sin_sub} sin sucursal en amarillo)")
