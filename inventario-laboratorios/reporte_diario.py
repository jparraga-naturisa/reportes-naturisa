"""
Reporte Diario Automatico - Naturisa
API: gateway.naturisa.com.ec
Reporte general + Alerta + Envio por correo Gmail
"""

import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os, sys, json
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

# -------------------------------------------------------------
# CONFIGURACION
# -------------------------------------------------------------
LOGIN_URL        = "https://gateway.naturisa.com.ec/bff/web/ap1/security/api/auth"
USUARIO          = os.environ.get("NATURISA_USUARIO", "jparraga")
PASSWORD         = os.environ.get("NATURISA_PASSWORD", "Naturisa2025")
CODE_APPLICATION = "55ab9cb4-c887-4f42-98ec-b90470be6613"

SUBSIDIARY_IDS = [13,28,29,30,19,6,8,7,11,14,17,16,15,5,4,18,21,3,10,9,1,33,2,12,20,10033]
BUSINESS_TYPES = ["WallIncome","WallBalance","HopperBalance","Remaining","Loaded","VoleoConsumption"]
OUTPUT_DIR     = os.environ.get("OUTPUT_DIR", r"C:\Reportes\Naturisa")

CORREO_REMITENTE    = "parragajonathan965@gmail.com"
PASSWORD_CORREO     = os.environ.get("GMAIL_PASSWORD", "wdiw eaif jwov bfxf")
CORREO_DESTINATARIO = ["jparraga@naturisa.com.ec", "asanlucas@naturisa.com.ec", "jcorozo@naturisa.com.ec", "rmaspons@naturisa.com.ec", "jlafuente@naturisa.com.ec"]
SMTP_SERVER         = "smtp.gmail.com"
SMTP_PORT           = 587
# -------------------------------------------------------------

BASE    = "https://gateway.naturisa.com.ec"
API_URL = f"{BASE}/bff/mobile/feedcontrol/balanceado/api/report/feeding_general"

NOMBRES = {
    "Loaded":           "Cargado Tolva",
    "VoleoConsumption": "Consumo Voleo",
    "WallIncome":       "Ingreso a Muro",
    "WallBalance":      "Saldo Muro",
    "HopperBalance":    "Saldo Tolva",
    "Remaining":        "Sobrante Tolva",
    "Cargado Tolva":    "Cargado Tolva",
    "Consumo Voleo":    "Consumo Voleo",
    "Ingreso a Muro":   "Ingreso a Muro",
    "Saldo Muro":       "Saldo Muro",
    "Saldo Tolva":      "Saldo Tolva",
    "Sobrante Tolva":   "Sobrante Tolva",
}

COLUMNAS_ORDEN = ["Ingreso a Muro","Saldo Muro","Cargado Tolva","Sobrante Tolva","Saldo Tolva"]


def obtener_token() -> str:
    try:
        r = requests.post(LOGIN_URL, json={
            "userName":        USUARIO,
            "password":        PASSWORD,
            "codeApplication": CODE_APPLICATION,
            "includeUserInfo": True
        }, timeout=15)
        r.raise_for_status()
        data = r.json()

        token = (data.get("token") or data.get("accessToken") or
                 data.get("access_token") or data.get("jwt") or
                 data.get("bearerToken"))

        if not token and isinstance(data, dict):
            for v in data.values():
                if isinstance(v, dict):
                    token = (v.get("token") or v.get("accessToken") or
                             v.get("access_token") or v.get("jwt"))
                    if token:
                        break

        if token:
            print(f"[{datetime.now():%H:%M:%S}] Login OK")
            return token

        print("Login OK pero no se encontro el token.")
        sys.exit(1)

    except Exception as e:
        print(f"Login fallo: {e}")
        sys.exit(1)


def obtener_datos(fecha: str, token: str) -> list:
    qs_ids   = "&".join(f"subsidiaryIds={i}" for i in SUBSIDIARY_IDS)
    qs_types = "&".join(f"businessTypes={b}" for b in BUSINESS_TYPES)
    url = (f"{API_URL}?{qs_ids}&{qs_types}"
           f"&initDate={fecha}&endDate={fecha}"
           f"&timeGranularity=day&valueOptions=kg&groupBy=pool&PageSize=1000")

    headers = {"Authorization": f"Bearer {token}", "Accept": "application/json"}
    print(f"[{datetime.now():%H:%M:%S}] Consultando API...")
    r = requests.get(url, headers=headers, timeout=30)
    r.raise_for_status()
    data = r.json()

    if isinstance(data, list):
        return data
    for key in ("data","records","items","result","content"):
        if key in data and isinstance(data[key], list):
            return data[key]

    with open("respuesta_debug.json","w") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    return []


def construir_tabla(registros: list) -> pd.DataFrame:
    rows = []
    for rec in registros:
        bt    = rec.get("businessType", "")
        psc   = rec.get("groupValue", "")
        sacos = rec.get("sacks", 0) or 0
        rows.append({"PSC": psc, "Metrica": NOMBRES.get(bt, bt), "Sacos": sacos})

    if not rows:
        return pd.DataFrame()

    df = pd.DataFrame(rows)
    tabla = df.pivot_table(index="PSC", columns="Metrica", values="Sacos",
                           aggfunc="sum", fill_value=0).reset_index()
    tabla.columns.name = None

    for col in COLUMNAS_ORDEN:
        if col not in tabla.columns:
            tabla[col] = 0

    cols_presentes = [c for c in COLUMNAS_ORDEN if c in tabla.columns]
    extra = [c for c in tabla.columns if c not in ["PSC"] + COLUMNAS_ORDEN]
    tabla = tabla[["PSC"] + cols_presentes + extra]
    return tabla.sort_values("PSC").reset_index(drop=True)


def detectar_alertas(df: pd.DataFrame) -> pd.DataFrame:
    for col in ["Saldo Tolva", "Cargado Tolva", "Sobrante Tolva"]:
        if col not in df.columns:
            df[col] = 0

    alertas = df[
        (df["Saldo Tolva"]    > 0) &
        (df["Cargado Tolva"]  == 0) &
        (df["Sobrante Tolva"] == 0)
    ].copy()
    return alertas.reset_index(drop=True)


def borde_celda():
    return Border(
        left=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
        top=Side(style="thin", color="BFBFBF"),
        bottom=Side(style="thin", color="BFBFBF"),
    )


def crear_excel_reporte(df: pd.DataFrame, fecha: str) -> str:
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    ruta = os.path.join(OUTPUT_DIR, f"Reporte_Naturisa_{fecha}.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Diario"
    borde = borde_celda()
    ult_letra = get_column_letter(len(df.columns))

    ws.merge_cells(f"A1:{ult_letra}1")
    ws["A1"] = f"Reporte General Naturisa - {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A1"].font      = Font(bold=True, color="FFFFFF", size=13, name="Arial")
    ws["A1"].fill      = PatternFill("solid", fgColor="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    for ci, nombre in enumerate(df.columns, 1):
        c = ws.cell(row=2, column=ci, value=nombre)
        c.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        c.fill = PatternFill("solid", fgColor="2E75B6")
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = borde
    ws.row_dimensions[2].height = 22

    for ri, row in enumerate(df.itertuples(index=False), 3):
        fill = PatternFill("solid", fgColor="D6E4F0" if ri % 2 == 0 else "F2F2F2")
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = Font(name="Arial", size=10)
            c.fill = fill
            c.border = borde
            c.alignment = Alignment(horizontal="left" if ci == 1 else "center")

    ult_dato = 2 + len(df)
    fila_tot = ult_dato + 1
    for ci, col_name in enumerate(df.columns, 1):
        letra = get_column_letter(ci)
        c = ws.cell(row=fila_tot, column=ci)
        c.fill = PatternFill("solid", fgColor="1F4E79")
        c.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        c.border = borde
        c.alignment = Alignment(horizontal="center")
        if col_name == "PSC":
            c.value = "TOTAL"
        elif pd.api.types.is_numeric_dtype(df[col_name]):
            c.value = f"=SUM({letra}3:{letra}{ult_dato})"

    for ci, col_name in enumerate(df.columns, 1):
        ws.column_dimensions[get_column_letter(ci)].width = max(len(col_name) + 2, 14)

    ws.freeze_panes = "B3"
    wb.save(ruta)
    return ruta


def crear_excel_alertas(df_alertas: pd.DataFrame, fecha: str) -> str:
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    ruta = os.path.join(OUTPUT_DIR, f"ALERTA_Naturisa_{fecha}.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Alertas"
    borde = borde_celda()
    columnas = list(df_alertas.columns)
    ult_letra = get_column_letter(len(columnas))

    ws.merge_cells(f"A1:{ult_letra}1")
    ws["A1"] = f"ALERTA - Piscinas con Saldo Tolva sin movimiento - {fecha}"
    ws["A1"].font      = Font(bold=True, color="FFFFFF", size=13, name="Arial")
    ws["A1"].fill      = PatternFill("solid", fgColor="8B1A1A")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells(f"A2:{ult_letra}2")
    ws["A2"] = "Criterio: Saldo Tolva > 0  |  Cargado Tolva = 0  |  Sobrante Tolva = 0"
    ws["A2"].font      = Font(italic=True, color="8B1A1A", size=10, name="Arial")
    ws["A2"].fill      = PatternFill("solid", fgColor="FDFEFE")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    for ci, nombre in enumerate(columnas, 1):
        c = ws.cell(row=3, column=ci, value=nombre)
        c.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        c.fill = PatternFill("solid", fgColor="C0392B")
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = borde
    ws.row_dimensions[3].height = 22

    for ri, row in enumerate(df_alertas.itertuples(index=False), 4):
        fill = PatternFill("solid", fgColor="F9EBEA" if ri % 2 == 0 else "FFFFFF")
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = Font(name="Arial", size=10, bold=(columnas[ci-1] == "Saldo Tolva"))
            c.fill = fill
            c.border = borde
            c.alignment = Alignment(horizontal="left" if ci == 1 else "center")

    fila_res = 4 + len(df_alertas) + 1
    ws.merge_cells(f"A{fila_res}:{ult_letra}{fila_res}")
    ws[f"A{fila_res}"] = f"Total piscinas en alerta: {len(df_alertas)}"
    ws[f"A{fila_res}"].font      = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    ws[f"A{fila_res}"].fill      = PatternFill("solid", fgColor="8B1A1A")
    ws[f"A{fila_res}"].alignment = Alignment(horizontal="center")

    for ci, col_name in enumerate(columnas, 1):
        ws.column_dimensions[get_column_letter(ci)].width = max(len(col_name) + 2, 14)

    ws.freeze_panes = "A4"
    wb.save(ruta)
    return ruta


def enviar_correo_ok(n_total: int, fecha: str):
    try:
        print(f"[{datetime.now():%H:%M:%S}] Enviando correo de confirmacion...")
        cuerpo = f"""
<div style="font-family:Arial, sans-serif; max-width:820px; margin:0;">

  <div style="background:#1F3864; padding:20px 24px; text-align:center;">
    <h2 style="color:#FFFFFF; margin:0; font-size:15px; letter-spacing:1.5px; font-weight:bold; text-transform:uppercase;">
      TODAS LAS PISCINAS CON MOVIMIENTO NORMAL
    </h2>
    <p style="color:#BDC3C7; margin:6px 0 0; font-size:12px;">
      Fecha: {fecha} &nbsp;&nbsp;|&nbsp;&nbsp; Piscinas revisadas: <b style="color:#FFFFFF;">{n_total}</b>
    </p>
  </div>

  <div style="padding:24px; border:1px solid #D0D8E4; background:#F4FBF4;">
    <p style="font-size:14px; color:#1a5c1a; font-weight:bold; margin:0 0 8px;">
      &#10003; Sin alertas para el dia de hoy
    </p>
    <p style="font-size:13px; color:#555; margin:0;">
      Se revisaron <b>{n_total}</b> piscinas. Ninguna presenta Saldo Tolva sin movimiento.<br>
      Todas las piscinas tienen actividad de carga o sobrante registrada.
    </p>
  </div>

  <p style="font-size:11px; color:#999; margin-top:12px; text-align:left;">
    Este correo fue generado automaticamente por el sistema de reportes Naturisa.
  </p>
</div>
        """

        msg = MIMEMultipart()
        msg["From"]    = CORREO_REMITENTE
        msg["To"]      = ", ".join(CORREO_DESTINATARIO)
        msg["Subject"] = f"AP1 PRO | Sin alertas - Todas las piscinas con movimiento normal - {fecha}"
        msg.attach(MIMEText(cuerpo, "html"))

        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as servidor:
            servidor.ehlo()
            servidor.starttls()
            servidor.login(CORREO_REMITENTE, PASSWORD_CORREO)
            servidor.sendmail(CORREO_REMITENTE, CORREO_DESTINATARIO, msg.as_string())

        print(f"[{datetime.now():%H:%M:%S}] Correo de confirmacion enviado a {CORREO_DESTINATARIO}")

    except Exception as e:
        print(f"Error al enviar correo de confirmacion: {e}")


def enviar_correo(df_alertas: pd.DataFrame, fecha: str):
    try:
        print(f"[{datetime.now():%H:%M:%S}] Enviando correo...")
        n_alertas = len(df_alertas)

        # Construir filas de la tabla
        filas_html = ""
        for i, (_, row) in enumerate(df_alertas.iterrows()):
            bg = "#EBF3FB" if i % 2 == 0 else "#FFFFFF"
            filas_html += f"""
            <tr style="background:{bg};">
                <td style="padding:8px 14px; border:1px solid #D0D8E4; font-weight:500;">{row['PSC']}</td>
                <td style="padding:8px 14px; border:1px solid #D0D8E4; text-align:center;">{int(row.get('Ingreso a Muro', 0))}</td>
                <td style="padding:8px 14px; border:1px solid #D0D8E4; text-align:center;">{int(row.get('Saldo Muro', 0))}</td>
                <td style="padding:8px 14px; border:1px solid #D0D8E4; text-align:center;">{int(row.get('Cargado Tolva', 0))}</td>
                <td style="padding:8px 14px; border:1px solid #D0D8E4; text-align:center;">{int(row.get('Sobrante Tolva', 0))}</td>
                <td style="padding:8px 14px; border:1px solid #D0D8E4; text-align:center;">{int(row.get('Consumo Voleo', 0))}</td>
                <td style="padding:8px 14px; border:1px solid #D0D8E4; text-align:center; font-weight:bold; color:#C0392B;">{int(row.get('Saldo Tolva', 0))}</td>
            </tr>"""

        cuerpo = f"""
<div style="font-family:Arial, sans-serif; max-width:820px; margin:0;">

  <div style="background:#1F3864; padding:20px 24px; text-align:center;">
    <h2 style="color:#FFFFFF; margin:0; font-size:15px; letter-spacing:1.5px; font-weight:bold; text-transform:uppercase;">
      PISCINAS CON SALDO TOLVA SIN MOVIMIENTO
    </h2>
    <p style="color:#BDC3C7; margin:6px 0 0; font-size:12px;">
      Fecha: {fecha} &nbsp;&nbsp;|&nbsp;&nbsp; Total piscinas en alerta: <b style="color:#FFFFFF;">{n_alertas}</b>
    </p>
  </div>

  <table style="width:100%; border-collapse:collapse; font-size:13px;">
    <thead>
      <tr style="background:#2E75B6; color:#FFFFFF;">
        <th style="padding:10px 14px; border:1px solid #1F5C9A; text-align:left; font-weight:bold;">PSC</th>
        <th style="padding:10px 14px; border:1px solid #1F5C9A; text-align:center; font-weight:bold;">Ingreso a Muro</th>
        <th style="padding:10px 14px; border:1px solid #1F5C9A; text-align:center; font-weight:bold;">Saldo Muro</th>
        <th style="padding:10px 14px; border:1px solid #1F5C9A; text-align:center; font-weight:bold;">Cargado Tolva</th>
        <th style="padding:10px 14px; border:1px solid #1F5C9A; text-align:center; font-weight:bold;">Sobrante Tolva</th>
        <th style="padding:10px 14px; border:1px solid #1F5C9A; text-align:center; font-weight:bold;">Consumo Voleo</th>
        <th style="padding:10px 14px; border:1px solid #1F5C9A; text-align:center; font-weight:bold;">Saldo Tolva</th>
      </tr>
    </thead>
    <tbody>
      {filas_html}
    </tbody>
  </table>

  <p style="font-size:11px; color:#999; margin-top:12px; text-align:left;">
    Este correo fue generado automaticamente por el sistema de reportes Naturisa.
  </p>
</div>
        """

        msg = MIMEMultipart()
        msg["From"]    = CORREO_REMITENTE
        msg["To"]      = ", ".join(CORREO_DESTINATARIO)
        msg["Subject"] = f"AP1 PRO | Piscinas con Saldo Tolva sin movimiento - {fecha}"
        msg.attach(MIMEText(cuerpo, "html"))

        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as servidor:
            servidor.ehlo()
            servidor.starttls()
            servidor.login(CORREO_REMITENTE, PASSWORD_CORREO)
            servidor.sendmail(CORREO_REMITENTE, CORREO_DESTINATARIO, msg.as_string())

        print(f"[{datetime.now():%H:%M:%S}] Correo enviado a {CORREO_DESTINATARIO}")

    except Exception as e:
        print(f"Error al enviar correo: {e}")


def main():
    from datetime import timezone, timedelta
    ECUADOR = timezone(timedelta(hours=-5))
    fecha = datetime.now(ECUADOR).strftime("%Y-%m-%d")
    print(f"[{datetime.now():%H:%M:%S}] Reporte Naturisa - {fecha}")

    try:
        token     = obtener_token()
        registros = obtener_datos(fecha, token)

        if not registros:
            print("Sin datos para hoy.")
            sys.exit(0)

        print(f"[{datetime.now():%H:%M:%S}] {len(registros)} registros recibidos")
        df = construir_tabla(registros)

        # Reporte general
        ruta_reporte = crear_excel_reporte(df, fecha)
        print(f"[{datetime.now():%H:%M:%S}] Reporte generado: {ruta_reporte}")

        # Alertas
        df_alertas = detectar_alertas(df)
        if df_alertas.empty:
            print(f"[{datetime.now():%H:%M:%S}] Sin alertas - todas las piscinas con movimiento normal")
            enviar_correo_ok(len(df), fecha)
        else:
            print(f"[{datetime.now():%H:%M:%S}] ALERTA: {len(df_alertas)} piscinas sin movimiento")
            crear_excel_alertas(df_alertas, fecha)
            enviar_correo(df_alertas, fecha)

    except requests.HTTPError as e:
        print(f"Error HTTP {e.response.status_code}: {e.response.text[:300]}")
        sys.exit(1)
    except Exception as e:
        print(f"Error: {e}")
        raise


if __name__ == "__main__":
    main()