"""
Reporte Saldos de Inventario - Naturisa Labs
Genera Excel por cada Larviquest con ciclo activo
Formato: Material | Descripcion | U/M | Stock SAP | Stock Lab | Consumos no Enviados | Diferencias
"""

import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, date
import os, sys, json
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

# -------------------------------------------------------------
# CONFIGURACION
# -------------------------------------------------------------
LOGIN_URL        = "https://gateway.naturisa.com.ec/bff/web/ap1/security/api/auth"
USUARIO          = "jparraga"
PASSWORD         = "Naturisa2025"
CODE_APPLICATION = "55ab9cb4-c887-4f42-98ec-b90470be6613"

SUBSIDIARIAS = {
    23: "Larviquest 1",
    24: "Larviquest 2",
    25: "Larviquest 3",
    26: "Larviquest 4",
}

# Código de almacén SAP por subsidiaria
ALMACEN_POR_SUBSIDIARIA = {
    23: "L109",
    24: "L209",
    25: "L309",
    26: "L409",
}

OUTPUT_DIR = r"C:\Reportes\Naturisa"

CORREO_REMITENTE    = "parragajonathan965@gmail.com"
PASSWORD_CORREO     = "wdiw eaif jwov bfxf"
CORREO_DESTINATARIO = ["jparraga@naturisa.com.ec" , "asanlucas@naturisa.com.ec", "jvillavicencio@naturisa.com.ec" , "jlafuente@naturisa.com.ec", "rmaspons@naturisa.com.ec" , "csilva@naturisa.com.ec"  , "larreaga@naturisa.com.ec" ]
SMTP_SERVER         = "smtp.gmail.com"
SMTP_PORT           = 587
# -------------------------------------------------------------

BASE     = "https://gateway.naturisa.com.ec"
LAB_BASE = f"{BASE}/bff/web/lab/backoffice/api"

DESKTOP = os.path.join(os.path.expanduser("~"), "Desktop")


def cargar_saldos_sap(ruta_excel: str = None) -> dict:
    """
    Lee el EXPORT_*.xlsx de SAP del Escritorio.
    Devuelve {almacen: {material: stock_libre}} separado por almacén.
    Columnas usadas: 'Material', 'Almacén', 'Libre utilización'.
    """
    if not ruta_excel:
        archivos = sorted(
            [f for f in os.listdir(DESKTOP) if f.upper().startswith("EXPORT_") and f.lower().endswith(".xlsx")],
            reverse=True
        )
        if not archivos:
            print("  ADVERTENCIA: No se encontro archivo EXPORT_*.xlsx en el Escritorio. Stock SAP sera estimado.")
            return {}
        ruta_excel = os.path.join(DESKTOP, archivos[0])

    print(f"  Leyendo saldos SAP desde: {os.path.basename(ruta_excel)}")
    try:
        df = pd.read_excel(ruta_excel, header=0, dtype=str)
        cols = {str(c).strip(): c for c in df.columns}

        col_mat = next((cols[c] for c in cols if c.lower() == "material"), None)
        col_alm = next((cols[c] for c in cols if "almac" in c.lower() and "denom" not in c.lower()), None)
        col_lib = next((cols[c] for c in cols if "libre" in c.lower() and "utiliz" in c.lower()), None)

        if not col_mat or not col_lib:
            print(f"  ADVERTENCIA: columnas no reconocidas. Disponibles: {list(cols.keys())}")
            return {}

        col_um  = next((cols[c] for c in cols if "unidad" in c.lower() and "medida" in c.lower()), None)
        col_desc = next((cols[c] for c in cols if "descripci" in c.lower() and "material" in c.lower()), None)

        FACTOR = {"KG": 1000, "KGS": 1000, "G": 1, "GR": 1,
                  "L": 1000, "LT": 1000, "ML": 1, "GAL": 3785.41, "GLN": 3785.41}
        UM_LAB = {"KG": "G", "KGS": "G", "G": "G", "GR": "G",
                  "L": "ML", "LT": "ML", "ML": "ML", "GAL": "ML", "GLN": "ML"}

        cols_sel = [c for c in [col_mat, col_alm, col_desc, col_um, col_lib] if c]
        df = df[cols_sel].copy()
        df[col_mat] = df[col_mat].astype(str).str.strip()
        df[col_lib] = pd.to_numeric(df[col_lib], errors="coerce").fillna(0)

        if col_um:
            df[col_um] = df[col_um].astype(str).str.strip().str.upper()
            df[col_lib] = df.apply(
                lambda r: round(r[col_lib] * FACTOR.get(r[col_um], 1), 4), axis=1
            )
            df["_um_lab"] = df[col_um].map(lambda u: UM_LAB.get(u, u))
        else:
            df["_um_lab"] = ""

        # Agrupar por almacén → {material: {stock, descripcion, unidad}}
        resultado = {}
        grupos = df.groupby(col_alm) if col_alm else [("__todos__", df)]
        for almacen, grupo in grupos:
            almacen = str(almacen).strip()
            materiales = {}
            for mat, filas in grupo.groupby(col_mat):
                if not mat or mat.lower() == "nan":
                    continue
                materiales[mat] = {
                    "stock":      round(filas[col_lib].sum(), 4),
                    "descripcion": filas[col_desc].iloc[0].strip() if col_desc else "",
                    "unidad":     filas["_um_lab"].iloc[0],
                }
            resultado[almacen] = materiales
            print(f"    {almacen}: {len(materiales)} materiales")

        return resultado
    except Exception as e:
        print(f"  ERROR leyendo Excel SAP: {e}")
        return {}


def obtener_token() -> str:
    try:
        r = requests.post(LOGIN_URL, json={
            "userName":        USUARIO,
            "password":        PASSWORD,
            "codeApplication": CODE_APPLICATION,
            "includeUserInfo": True
        }, timeout=15)
        r.raise_for_status()
        data = r.json()
        token = (data.get("token") or data.get("accessToken") or
                 data.get("access_token") or data.get("jwt") or
                 data.get("bearerToken"))
        if not token and isinstance(data, dict):
            for v in data.values():
                if isinstance(v, dict):
                    token = (v.get("token") or v.get("accessToken") or
                             v.get("access_token") or v.get("jwt"))
                    if token:
                        break
        if token:
            print(f"[{datetime.now():%H:%M:%S}] Login OK")
            return token
        print("No se encontro el token.")
        sys.exit(1)
    except Exception as e:
        print(f"Login fallo: {e}")
        sys.exit(1)


def obtener_ciclo_activo(token: str, subsidiary_id: int) -> dict:
    url = (f"{LAB_BASE}/cycle?PageNumber=1&PageSize=1000&OrderBy=code+desc"
           f"&subsidiaryId={subsidiary_id}&includeLaboratory=true&includeLarvalStage=true")
    headers = {"Authorization": f"Bearer {token}", "Accept": "application/json"}
    r = requests.get(url, headers=headers, timeout=15)
    r.raise_for_status()
    data = r.json()

    ciclos = []
    if isinstance(data, list):
        ciclos = data
    elif "data" in data:
        inner = data["data"]
        if isinstance(inner, list):
            ciclos = inner
        elif isinstance(inner, dict) and "data" in inner:
            ciclos = inner["data"]

    if ciclos:
        ciclo    = ciclos[0]
        cycle_id = ciclo.get("id") or ciclo.get("cycleId") or ciclo.get("idCycle")
        fecha_inicio = ciclo.get("sowingDate") or ciclo.get("dryDate") or date.today().strftime("%Y-%m-%d")
        if "T" in str(fecha_inicio):
            fecha_inicio = fecha_inicio.split("T")[0]
        print(f"  Ciclo ID: {cycle_id} | Fecha inicio: {fecha_inicio}")
        return {"cycle_id": cycle_id, "fecha_inicio": fecha_inicio}

    print(f"  Sin ciclo activo para subsidiaria {subsidiary_id}")
    return None


def obtener_inventario(token: str, subsidiary_id: int, cycle_id: int, fecha_inicio: str) -> list:
    headers = {"Authorization": f"Bearer {token}", "Accept": "application/json"}
    todos   = []
    page    = 1

    while True:
        url = (f"{LAB_BASE}/warehouse_production"
               f"?subsidiaryId={subsidiary_id}"
               f"&cycleId={cycle_id}"
               f"&date={fecha_inicio}"
               f"&includeItem=true"
               f"&includeWarehouseOrigin=true"
               f"&PageSize=100"
               f"&PageNumber={page}")

        r = requests.get(url, headers=headers, timeout=30)
        r.raise_for_status()
        data = r.json()

        registros   = []
        total_pages = 1

        if isinstance(data, list):
            registros = data
        elif "data" in data:
            inner = data["data"]
            if isinstance(inner, list):
                registros = inner
            elif isinstance(inner, dict):
                registros   = inner.get("data", [])
                total_pages = inner.get("totalPages", 1)

        todos.extend(registros)
        if page >= total_pages or not registros:
            break
        page += 1

    return todos


def construir_tabla(registros: list, saldos_sap: dict) -> pd.DataFrame:
    """
    Construye la tabla combinando datos del API (lab) y del Excel SAP.
    Base: todos los materiales que aparecen en SAP.
    Si el material también está en el API, se completan Stock Lab y Consumos.
    Si solo está en SAP, Stock Lab = 0 y Consumos = 0.
    Si solo está en el API (no en SAP), Stock SAP = Stock Lab + Consumos (estimado).
    """
    UM_LAB = {"KG": "G", "KGS": "G", "G": "G", "GR": "G",
              "L": "ML", "LT": "ML", "ML": "ML", "GAL": "ML", "GLN": "ML"}

    # Indexar registros del API por código SAP
    api_por_codigo = {}
    for rec in registros:
        item   = rec.get("item", {})
        codigo = str(item.get("codeSap", "")).strip()
        if codigo:
            api_por_codigo[codigo] = rec

    rows = []

    # 1. Todos los materiales del Excel SAP
    for codigo, info_sap in saldos_sap.items():
        rec            = api_por_codigo.get(codigo, {})
        item           = rec.get("item", {}) if rec else {}
        saldo_lab      = rec.get("stockAvailable", 0) or 0 if rec else 0
        consumo_nosync = rec.get("unsyncQuantity", 0) or 0 if rec else 0
        stock_sap      = info_sap["stock"]

        unidad_api = item.get("baseUnitMsrAb", "").upper()
        unidad_lab = UM_LAB.get(unidad_api) or info_sap["unidad"] or unidad_api or ""

        descripcion = item.get("name", "") or info_sap["descripcion"]

        rows.append({
            "Material":             codigo,
            "Descripcion":          descripcion,
            "Unidad Medida":        unidad_lab,
            "Stock SAP":            round(stock_sap, 2),
            "Stock Laboratorio":    round(saldo_lab, 2),
            "Consumos no Enviados": round(consumo_nosync, 2),
            "Diferencias":          round(stock_sap - saldo_lab - consumo_nosync, 2),
        })

    # 2. Materiales del API que no están en el Excel SAP
    for codigo, rec in api_por_codigo.items():
        if codigo in saldos_sap:
            continue
        item           = rec.get("item", {})
        saldo_lab      = rec.get("stockAvailable", 0) or 0
        consumo_nosync = rec.get("unsyncQuantity", 0) or 0
        stock_sap      = saldo_lab + consumo_nosync  # estimado

        unidad_api = item.get("baseUnitMsrAb", "").upper()
        unidad_lab = UM_LAB.get(unidad_api, unidad_api)

        rows.append({
            "Material":             codigo,
            "Descripcion":          item.get("name", ""),
            "Unidad Medida":        unidad_lab,
            "Stock SAP":            round(stock_sap, 2),
            "Stock Laboratorio":    round(saldo_lab, 2),
            "Consumos no Enviados": round(consumo_nosync, 2),
            "Diferencias":          0.0,
        })

    return pd.DataFrame(rows).sort_values("Material").reset_index(drop=True)


def borde_celda():
    return Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )


def crear_excel(df: pd.DataFrame, fecha: str, nombre_sub: str) -> str:
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    nombre_archivo = nombre_sub.replace(" ", "_")
    ruta = os.path.join(OUTPUT_DIR, f"Inventario_{nombre_archivo}_{fecha}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Saldos Inventario"
    borde = borde_celda()

    GRIS_OSC  = "2F2F2F"
    GRIS_MED  = "595959"
    AZUL_CLAR = "EBF3FB"
    BLANCO    = "FFFFFF"
    ROJO_CLAR = "FFE0E0"
    ROJO_FONT = "C00000"

    columnas  = list(df.columns)
    n_cols    = len(columnas)
    ult_letra = get_column_letter(n_cols)

    # Titulo
    ws.merge_cells(f"A1:{ult_letra}1")
    ws["A1"] = f"Saldos de Inventario — {nombre_sub} — {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A1"].font      = Font(bold=True, color="FFFFFF", size=12, name="Arial")
    ws["A1"].fill      = PatternFill("solid", fgColor=GRIS_OSC)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # Encabezados — nombres como en el Excel de referencia
    encabezados = {
        "Material":             "Material",
        "Descripcion":          "Descripción del material",
        "Unidad Medida":        "Unidad\nMedida",
        "Stock SAP":            "Stock\nSAP",
        "Stock Laboratorio":    "Stock\nLaboratorio",
        "Consumos no Enviados": "Consumos no\nEnviados",
        "Diferencias":          "Diferencias",
    }
    for ci, col in enumerate(columnas, 1):
        c = ws.cell(row=2, column=ci, value=encabezados.get(col, col))
        c.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        c.fill      = PatternFill("solid", fgColor=GRIS_MED)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = borde
    ws.row_dimensions[2].height = 30

    # Datos
    col_dif_idx = columnas.index("Diferencias") + 1
    for ri, row in enumerate(df.itertuples(index=False), 3):
        dif_val     = row[col_dif_idx - 1]
        fill_normal = PatternFill("solid", fgColor=AZUL_CLAR if ri % 2 == 0 else BLANCO)
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = borde
            if ci == col_dif_idx and dif_val != 0:
                c.fill = PatternFill("solid", fgColor=ROJO_CLAR)
                c.font = Font(name="Arial", size=10, bold=True, color=ROJO_FONT)
            else:
                c.fill = fill_normal
                c.font = Font(name="Arial", size=10)
            if ci <= 2:
                c.alignment = Alignment(horizontal="left", vertical="center")
            else:
                c.alignment = Alignment(horizontal="center", vertical="center")
                if ci > 3 and isinstance(val, (int, float)):
                    c.number_format = "#,##0.00"

    # Totales
    ult_dato = 2 + len(df)
    fila_tot = ult_dato + 1
    cols_sum = ["Stock SAP","Stock Laboratorio","Consumos no Enviados","Diferencias"]
    for ci, col_name in enumerate(columnas, 1):
        letra = get_column_letter(ci)
        c = ws.cell(row=fila_tot, column=ci)
        c.border = borde
        c.alignment = Alignment(horizontal="center", vertical="center")
        if col_name == "Material":
            c.value = "TOTAL"
            c.font  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
            c.fill  = PatternFill("solid", fgColor=GRIS_OSC)
        elif col_name in cols_sum:
            c.value = f"=SUM({letra}3:{letra}{ult_dato})"
            c.number_format = "#,##0.00"
            c.font  = Font(bold=True, color=ROJO_FONT if col_name == "Diferencias" else "FFFFFF",
                          name="Arial", size=10)
            c.fill  = PatternFill("solid", fgColor=GRIS_OSC)
        else:
            c.fill = PatternFill("solid", fgColor=GRIS_OSC)
            c.font = Font(name="Arial", size=10, color="FFFFFF")

    # Anchos
    anchos = {
        "Material": 12, "Descripcion": 45, "Unidad Medida": 10,
        "Stock SAP": 14, "Stock Laboratorio": 16,
        "Consumos no Enviados": 18, "Diferencias": 14
    }
    for ci, col_name in enumerate(columnas, 1):
        ws.column_dimensions[get_column_letter(ci)].width = anchos.get(col_name, 14)

    ws.freeze_panes = "D3"
    wb.save(ruta)
    return ruta


def tabla_html(df, nombre_sub: str) -> str:
    """Genera tabla HTML de un laboratorio. df puede ser None si no hay produccion."""
    abrev = nombre_sub.replace("Larviquest ", "LQ")

    TD  = "padding:10px 14px; border:1px solid #ddd; font-size:14px; font-weight:bold;"
    TDR = TD + " text-align:right;"
    TDC = TD + " text-align:center;"

    if df is None:
        return f"""
    <div style="margin-bottom:28px;">
      <div style="background:#2F2F2F; padding:12px 18px; border-radius:4px 4px 0 0; text-align:center;">
        <b style="color:#FFFFFF; font-size:17px; letter-spacing:1px;">{abrev}</b>
      </div>
      <div style="border:1px solid #ddd; border-top:none; padding:16px; background:#FAFAFA; border-radius:0 0 4px 4px; text-align:center;">
        <p style="margin:0; font-size:14px; font-weight:bold; color:#888; font-style:italic;">
          &#128274; Ciclo en fase de secado — sin produccion activa
        </p>
      </div>
    </div>"""

    filas = ""
    for i, (_, row) in enumerate(df.iterrows()):
        bg  = "#F0F4F8" if i % 2 == 0 else "#FFFFFF"
        dif = row["Diferencias"]
        td_dif = (TD + " color:#C00000;") if dif != 0 else TD
        filas += f"""
        <tr style="background:{bg};">
            <td style="{TD}">{row["Material"]}</td>
            <td style="{TD} white-space:nowrap;">{row["Descripcion"]}</td>
            <td style="{TDC}">{row["Unidad Medida"]}</td>
            <td style="{TDR}">{row["Stock SAP"]:,.2f}</td>
            <td style="{TDR}">{row["Stock Laboratorio"]:,.2f}</td>
            <td style="{TDR}">{row["Consumos no Enviados"]:,.2f}</td>
            <td style="{td_dif} text-align:right;">{dif:,.2f}</td>
        </tr>"""

    TT = "padding:11px 14px; border:1px solid #163a5f; font-size:15px; font-weight:bold; text-align:right; background:#1F4E79; color:#FFFFFF;"
    filas += f"""
        <tr>
            <td style="{TT} text-align:center; min-width:110px;">TOTAL</td>
            <td style="{TT} min-width:320px;"></td>
            <td style="{TT} min-width:55px;"></td>
            <td style="{TT} min-width:110px;">{df["Stock SAP"].sum():,.2f}</td>
            <td style="{TT} min-width:140px;">{df["Stock Laboratorio"].sum():,.2f}</td>
            <td style="{TT} min-width:160px;">{df["Consumos no Enviados"].sum():,.2f}</td>
            <td style="{TT} min-width:110px; color:#FFD700;">{df["Diferencias"].sum():,.2f}</td>
        </tr>"""

    TH = "padding:11px 14px; border:1px solid #1a5c9e; font-size:14px; text-align:center;"
    return f"""
    <div style="margin-bottom:32px;">
      <div style="background:#1F4E79; padding:14px 18px; border-radius:4px 4px 0 0; text-align:center;">
        <b style="color:#FFFFFF; font-size:18px; letter-spacing:2px;">{abrev}</b>
      </div>
      <table style="width:100%; border-collapse:collapse; border:1px solid #ddd; border-top:none; table-layout:fixed;">
        <colgroup>
          <col style="width:11%;">
          <col style="width:33%;">
          <col style="width:5%;">
          <col style="width:12%;">
          <col style="width:14%;">
          <col style="width:14%;">
          <col style="width:11%;">
        </colgroup>
        <thead>
          <tr style="background:#2E75B6; color:#FFFFFF;">
            <th style="{TH}">Material</th>
            <th style="{TH}">Descripción del material</th>
            <th style="{TH}">UM</th>
            <th style="{TH}">Stock SAP</th>
            <th style="{TH}">Stock Laboratorio</th>
            <th style="{TH}">Consumos no Enviados</th>
            <th style="{TH}">Diferencias</th>
          </tr>
        </thead>
        <tbody>{filas}</tbody>
      </table>
    </div>"""


def enviar_correo(tablas_data: list, fecha: str):
    """Envia correo con tablas en el cuerpo del correo."""
    try:
        print(f"[{datetime.now():%H:%M:%S}] Enviando correo...")
        msg = MIMEMultipart()
        msg["From"]    = CORREO_REMITENTE
        msg["To"]      = ", ".join(CORREO_DESTINATARIO)
        msg["Subject"] = f"Control de Bodegas de Produccion Laboratorios — {fecha}"

        tablas_html = "".join(tabla_html(df, nombre) for df, nombre in tablas_data)

        cuerpo = f"""
<div style="font-family:Arial, sans-serif; width:100%;">
  <div style="background:#1a1a2e; padding:22px 24px; border-radius:6px 6px 0 0; margin-bottom:24px; text-align:center;">
    <h2 style="color:#FFFFFF; margin:0; font-size:20px; font-weight:bold; letter-spacing:2px;">
      CONTROL DE BODEGAS DE PRODUCCION LABORATORIOS
    </h2>
    <p style="color:#8AB8D8; margin:8px 0 0; font-size:14px; font-weight:bold;">Fecha: {fecha}</p>
  </div>
  {tablas_html}
  <p style="font-size:12px; color:#999; margin-top:12px; text-align:center;">
    <i>Este correo fue generado automaticamente por el sistema de reportes Naturisa.</i>
  </p>
</div>"""

        msg.attach(MIMEText(cuerpo, "html"))

        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as servidor:
            servidor.ehlo()
            servidor.starttls()
            servidor.login(CORREO_REMITENTE, PASSWORD_CORREO)
            servidor.sendmail(CORREO_REMITENTE, CORREO_DESTINATARIO, msg.as_string())

        print(f"[{datetime.now():%H:%M:%S}] Correo enviado a {CORREO_DESTINATARIO}")

    except Exception as e:
        print(f"Error al enviar correo: {e}")


def main(ruta_excel: str = None):
    fecha = date.today().strftime("%Y-%m-%d")
    print(f"[{datetime.now():%H:%M:%S}] Reporte Inventario Naturisa — {fecha}")

    try:
        token = obtener_token()

        saldos_sap_por_almacen = cargar_saldos_sap(ruta_excel)

        rutas_generadas = []
        tablas_data = []

        for sub_id, sub_nombre in SUBSIDIARIAS.items():
            print(f"\n--- {sub_nombre} (ID: {sub_id}) ---")

            # Obtener saldos SAP del almacén correspondiente a esta subsidiaria
            almacen    = ALMACEN_POR_SUBSIDIARIA.get(sub_id, "")
            saldos_sap = (saldos_sap_por_almacen.get(almacen)
                          or saldos_sap_por_almacen.get("__todos__")
                          or {})
            if saldos_sap:
                print(f"  Almacen SAP: {almacen} ({len(saldos_sap)} materiales)")
            else:
                print(f"  ADVERTENCIA: sin datos SAP para almacen {almacen}")

            ciclo     = obtener_ciclo_activo(token, sub_id)
            registros = []
            if ciclo:
                registros = obtener_inventario(token, sub_id, ciclo["cycle_id"], ciclo["fecha_inicio"])

            sin_app = not ciclo or not registros

            if sin_app:
                # Sin datos en la app — verificar si hay stock SAP > 0
                sap_con_stock = {k: v for k, v in saldos_sap.items() if v["stock"] > 0}
                if not sap_con_stock:
                    motivo = "sin ciclo activo" if not ciclo else "ciclo en seco"
                    print(f"  {sub_nombre} — {motivo} y sin stock SAP. Se omite.")
                    tablas_data.append((None, sub_nombre))
                    continue
                print(f"  Sin datos en app, pero {len(sap_con_stock)} materiales con stock SAP > 0. Incluyendo.")

            print(f"  {len(registros)} registros app | {len(saldos_sap)} materiales SAP")
            df   = construir_tabla(registros, saldos_sap)
            ruta = crear_excel(df, fecha, sub_nombre)
            rutas_generadas.append(ruta)
            tablas_data.append((df, sub_nombre))
            print(f"  Reporte generado: {ruta}")

            con_diff = df[df["Diferencias"] != 0]
            if con_diff.empty:
                print(f"  Sin diferencias")
            else:
                print(f"  {len(con_diff)} productos con diferencia")

        print(f"\n[{datetime.now():%H:%M:%S}] Proceso completado.")

        # Siempre enviar correo con todos los laboratorios
        enviar_correo(tablas_data, fecha)

    except requests.HTTPError as e:
        print(f"Error HTTP {e.response.status_code}: {e.response.text[:300]}")
        sys.exit(1)
    except Exception as e:
        print(f"Error: {e}")
        raise


if __name__ == "__main__":
    main()